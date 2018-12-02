#! python3

# お試し
# まだPCにエクセルをインストールしていないため、
# とりあえず処理のイメージだけ書いておく
# エクセルをインストールしてから動かしてみる

import openpyxl
import os
import tarfile

# tarファイル内のログファイル名
FILELIST = [
    'apple.txt',
    'banana.txt']

INPUTDIR = 'log'
OUTPUTDIR = 'out'

for foldername, subfolders, filenames in os.walk(INPUTDIR):

    # tarファイル名を取得
    for filename in filenames:
        tf = tarfile.open(os.path.join(INPUTDIR, filename), 'r')

        # 新規エクセルファイルを作成
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        BASEDIR = filename.split('.')[0]

        for fn in FILELIST:
            # tarファイル内のログファイルオブジェクトを取得
            ti = tf.getmember(BASEDIR + '/' + fn)
            f = tf.extractfile(fi)

            # ログファイル名と同じシートを作成
            wb.create_sheet(title=fn)
            sheet = wb[fn]

            # エクセルの行数
            cnt = 1

            for i in f.readlines():
                # A列にログファイルの内容をコピー
                cnt = cnt + 1
                sheet.cell(row=cnt, column=1).value = i

        # エクセルファイルに書き込み
        OUTPUTFILE = 'SIT_' + BASEDIR + '.xlsx'
        wb.save(os.path.join(OUTPUTDIR, OUTPUTFILE))

